import os
import json
import base64
import imghdr
import hashlib
import zipfile
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import load_workbook
from kparser.common.types_utils import get_random_uuid
from kparser.parserground.parser.excel_to_images import excel_to_images_with_uno, upload_image_to_tos, is_image_file # 你已实现的sheet转图片
from openai import OpenAI
from kparser.common.log_utils import get_logger
from kparser.common.config import VISION

logger = get_logger(__name__)

# 从配置文件读取多模态API配置
vlm_api_key = VISION.get("vlm_api_key", "")
vlm_api_base = VISION.get("vlm_api_base", "https://open.bigmodel.cn/api/paas/v4")
vlm_model = VISION.get("vlm_model", "glm-4v-plus-0111")
excel_vision_prompt = VISION.get("excel_vision_prompt", "请仔细分析这张图片中的所有内容，包括所有表格和子图。要求：1. 完整提取所有可见的文字信息，包括表格内容、图表标题、图例等 2. 保持原有的层级结构和组织方式 3. 对于表格数据，确保表头和对应的数据正确对齐 4. 对于子图，提取其标题、说明文字和相关数据 5. 使用嵌套的 JSON 结构来组织信息 请确保输出的 JSON 格式正确，所有属性名和属性值都正确对应。")

client = OpenAI(api_key=vlm_api_key, base_url=vlm_api_base)


def extract_sub_image_from_excel(excel_path):
    """
    从Excel文件中提取所有图片并保存，避免重复
    
    Args:
        file_path: Excel文件路径
        output_dir: 图片输出目录
        resolution_threshold: 定义高分辨率的阈值，宽高均大于此值才会被分析
        
    Returns:
        sub_image_url_list: 子图片路径列表
    """
    sub_image_url_list = []
    sub_image_path_list = []
    seen_hashes = set()  # 跟踪已处理的图片哈希值
    img_cnt = 0
    try:
        # 方法1: 使用openpyxl提取标准图片
        workbook = load_workbook(excel_path)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if hasattr(worksheet, '_images') and worksheet._images:
                logger.info(f"在sheet '{sheet_name}' 中发现 {len(worksheet._images)} 张标准图片")
                for img in worksheet._images:
                    try:
                        img_data = img._data()
                        img_hash = hashlib.md5(img_data).hexdigest()[:8]
                        # 跳过已处理的图片
                        if img_hash in seen_hashes:
                            continue    
                        # 跳过非图片文件
                        if not is_image_file(img_data):
                            continue
                        seen_hashes.add(img_hash)
                        img_format = "png"
                        try:
                            with BytesIO(img_data) as img_stream:
                                pil_img = PILImage.open(img_stream)
                                img_format = pil_img.format.lower() if pil_img.format else "png"
                        except Exception:
                            continue  # 无法识别的图片格式
                        # 创建文件、保存并上传
                        img_path = f"tmp_{img_cnt}.{img_format}"
                        img_cnt += 1
                        sub_image_path_list.append(img_path)
                        with open(img_path, 'wb') as f:
                            f.write(img_data)
                        url = upload_image_to_tos(img_path)
                        sub_image_url_list.append(url)
                        logger.info(f"保存标准图片到TOS, url: {url}")
                    except Exception as e:
                        print(f"处理图片时出错: {e}")
                        continue
        workbook.close()
        # 方法2: 使用zip解压提取所有媒体文件
        logger.info("\n开始提取所有媒体文件...")
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            media_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
            if not media_files:
                logger.info("未找到任何媒体文件")
                return sub_image_url_list
            logger.info(f"发现 {len(media_files)} 个媒体文件")
            for media_file in media_files:
                try:
                    file_data = zip_ref.read(media_file)
                    img_hash = hashlib.md5(file_data).hexdigest()[:8]
                    # 跳过已处理的图片
                    if img_hash in seen_hashes:
                        continue
                    # 跳过非图片文件
                    if not is_image_file(file_data):
                        continue
                    seen_hashes.add(img_hash)
                    # 获取文件扩展名
                    ext = imghdr.what(None, h=file_data) or "png"
                    img_path = f"tmp_{img_cnt}.{ext}"
                    img_cnt += 1
                    sub_image_path_list.append(img_path)
                    # 创建文件、保存并上传
                    with open(img_path, 'wb') as f:
                        f.write(file_data)
                    url = upload_image_to_tos(img_path)
                    sub_image_url_list.append(url)
                    logger.info(f"保存媒体图片到TOS, url: {url}")
                except Exception as e:
                    logger.error(f"处理媒体文件 {media_file} 失败: {e}")
                    continue
        
        return sub_image_url_list, sub_image_path_list
        
    except Exception as e:
        logger.error(f"从Excel提取图片失败: {e}")
        return [], sub_image_path_list

def extract_image_content(image_path):
    try:
        logger.info(f"start extract_image_content, image_path: {image_path}")
        with open(image_path, 'rb') as img_file:
            img_base = base64.b64encode(img_file.read()).decode('utf-8')
        
        response = client.chat.completions.create(
            model=vlm_model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_base}"}},
                        {"type": "text", "text": excel_vision_prompt}
                    ]
                }
            ]
        )
        content = response.choices[0].message.content
        logger.info(f"end extract_image_content: {content}")
        return content
    except Exception as e:
        logger.error(f"Failed to extract image content: {e}")
        return ""


class RAGExcelVisionParser:
    def __call__(self, binary, parser_config):
        tmp_dir = "tmp/excel_vision"
        os.makedirs(tmp_dir, exist_ok=True)
        # 1. 保存二进制为临时excel文件
        excel_path = os.path.join(tmp_dir, get_random_uuid() + ".xlsx")
        with open(excel_path, "wb") as f:
            f.write(binary)
        # 2. 转图片
        image_paths = excel_to_images_with_uno(excel_path, tmp_dir)
        logger.info(f"image_paths: {image_paths}")
        results = []
        for idx, img_path in enumerate(image_paths):
            logger.info(f"idx: {idx}, img_path: {img_path}")
            try:
                # 3. 多模态抽取
                content_json = json.loads(extract_image_content(img_path)[7:-3])
                logger.info(f"content_json: {content_json}")                
                # 4. 上传图片到TOS
                img_url = upload_image_to_tos(img_path)
                logger.info(f"img_url: {img_url}")
                
                
                results.append({
                    "id": get_random_uuid(),
                    "content": content_json,
                    "content_type": "DICT" if type(content_json) == dict else "TEXT",
                    "page_idx": [idx + 1],
                    "extra_data": {
                        "img_url": img_url,
                    }
                })
            except Exception as e:
                logger.error(f"Error processing image {img_path}: {e}")
                # 即使处理失败也要删除本地文件
                try:
                    os.remove(img_path)
                    logger.info(f"Deleted local image files after error: {img_path}")
                except Exception as del_e:
                    logger.warning(f"Failed to delete local image file {img_path}: {del_e}")
                continue
        # 3. 删除临时excel文件
        try:
            os.remove(excel_path)
            logger.info(f"Deleted temporary excel file: {excel_path}")
        except Exception as e:
            logger.warning(f"Failed to delete temporary excel file {excel_path}: {e}")
        return results