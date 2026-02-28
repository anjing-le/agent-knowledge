import openpyxl
import xlrd
from io import BytesIO

from kparser.common.types_utils import get_random_uuid
from kparser.common.log_utils import get_logger
from kparser.parserground.parser.excel_to_images import extract_image_from_cell, upload_image_to_tos

logger = get_logger(__name__)


def parse_excel_row_format(data, method):
    """
    解析 Excel 并返回有实际数据的所有行内容。

    :param data: Excel 文件的二进制流
    :param method: "ROW_HEADER" 时，前缀加 header，否则仅输出值
    :return: list[dict]，每个元素是一个字典，key为列名，value为对应行元素，增加sheet名称
    """
    result = []

    # 读取 Excel 文件
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        is_xls = False
        logger.info("xlsx format")
    except Exception:
        wb = xlrd.open_workbook(file_contents=data)
        is_xls = True
        logger.info("xls format")

    for sheet in (wb.sheets() if is_xls else wb.worksheets):  # 兼容 .xls 和 .xlsx
        rows = []
        sheet_name = sheet.name if is_xls else sheet.title

        if is_xls:
            for row_idx in range(sheet.nrows):
                rows.append(sheet.row_values(row_idx))
        else:
            rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue  # 跳过空 sheet

        headers = rows[0]
        # 只保留header非空且非None的列索引
        valid_indices = [i for i, h in enumerate(headers) if h is not None and str(h).strip()]
        valid_headers = [str(headers[i]).strip() for i in valid_indices]
        has_header = len(valid_indices) > 0

        for row in rows[1:] if has_header else rows:
            row_dict = {valid_headers[idx]: (str(row[j]).strip() if j < len(row) and row[j] is not None else "") for idx, j in enumerate(valid_indices)}
            # 跳过完全为空的行（所有value都为""）
            if any([v != "" for v in row_dict.values()]):
                row_dict["sheet"] = sheet_name
                result.append(row_dict)

    return result


def parse_excel_first_col_group_by_column(data):
    """
    解析 Excel 文件，第一列作为"行标签"，其余列按列分组，生成多行字符串，
    所有工作表的内容合并到一个统一的列表中，忽略空值。

    :param data: Excel 文件的二进制流 (io.BytesIO)
    :return: 列表，其中包含每列按行标签: 数据格式分组后的字符串
    """
    result = []

    # 读取 Excel 文件流
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        is_xls = False
    except Exception:
        wb = xlrd.open_workbook(file_contents=data)
        is_xls = True

    for sheet in (wb.sheets() if is_xls else wb.worksheets):  # 兼容 .xls 和 .xlsx
        rows = []

        if is_xls:
            for row_idx in range(sheet.nrows):
                rows.append(sheet.row_values(row_idx))
        else:
            rows = list(sheet.iter_rows(values_only=True))

        # 跳过空的工作表或者只有一列的工作表
        if not rows or len(rows[0]) <= 1:
            continue

        # 初始化列输出
        column_outputs = ['' for _ in range(len(rows[0]) - 1)]  # 每列一个空字符串（不包括第一列）

        # 处理每一行
        for row in rows:
            row_header = str(row[0]).strip() if row[0] is not None else ""  # 第一列作为行标签
            if not row_header:
                continue  # 跳过空的行标签

            # 将行标签和列数据组合
            for j in range(1, len(row)):  # 从第二列开始处理
                cell_value = str(row[j]).strip() if row[j] is not None else ""  # 处理单元格的值
                if cell_value:  # 忽略空的单元格
                    column_outputs[j - 1] += f"{row_header}: {cell_value}\n"

        # 去掉全空列，并过滤掉无效内容
        filtered_outputs = [col.rstrip("\n") for col in column_outputs if col.strip()]
        result.extend(filtered_outputs)  # 将每个非空的列输出添加到结果列表中

    return result


def parse_excel_group_columns(data):
    """
    解析 Excel 文件，第一列作为内容的一部分，所有列按行分组输出，
    所有工作表的内容合并到一个统一的列表中，忽略空值。

    :param data: Excel 文件的二进制流 (io.BytesIO)
    :return: 列表，其中包含每列按数据格式分组后的字符串
    """
    result = []

    # 读取 Excel 文件流
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        is_xls = False
    except Exception:
        wb = xlrd.open_workbook(file_contents=data)
        is_xls = True

    for sheet in (wb.sheets() if is_xls else wb.worksheets):  # 兼容 .xls 和 .xlsx
        rows = []

        if is_xls:
            for row_idx in range(sheet.nrows):
                rows.append(sheet.row_values(row_idx))
        else:
            rows = list(sheet.iter_rows(values_only=True))

        # 跳过空的工作表或者只有一列的工作表
        if not rows or len(rows[0]) <= 1:
            continue

        # 初始化列输出
        column_outputs = ['' for _ in range(len(rows[0]))]  # 包括第一列

        # 处理每一行
        for row in rows:
            for j in range(len(row)):  # 从第一列到最后一列都处理
                cell_value = str(row[j]).strip() if row[j] is not None else ""  # 处理单元格的值
                if cell_value:  # 忽略空的单元格
                    column_outputs[j] += f"{cell_value}\n"  # 将列数据按行拼接

        # 去掉全空列，并过滤掉无效内容
        filtered_outputs = [col.rstrip("\n") for col in column_outputs if col.strip()]
        result.extend(filtered_outputs)  # 将每个非空的列输出添加到结果列表中

    return result


def analyse_table(data):
    cks_by_page = []
    for sec in data:
        element_text = {"id": get_random_uuid(),
                        "content": sec,
                        "content_type": "DICT" if type(sec) == dict else "TEXT",
                        "page_idx": [1],
                        "extra_data": {}
                        }
        # 保存文本
        cks_by_page.append(element_text)

    return cks_by_page


def get_rule_config(parser_rules):
    parser_rule = "ROW_HEADER"
    for item in parser_rules:
        if item["rule_method"] == "3":
            parser_rule = item["feature_value"][0]
    return parser_rule


class RAGExcelParser:
    def __call__(self, binary, parser_config):
        method = get_rule_config(parser_config["rules"])
        logger.info(f"EXCEL parser rule={method}")
        return self.load(binary, method)

    @classmethod
    def load(cls, data, method):
        # 调用函数解析 excel 数据，method 参数为 "COLUMN_HEADER" 或其他值
        if "ROW" in method:
            parsed_result = parse_excel_row_format(data, method)
        else:
            if method == "COLUMN_HEADER":
                parsed_result = parse_excel_first_col_group_by_column(data)
            else:
                parsed_result = parse_excel_group_columns(data)

        res = analyse_table(parsed_result)
        return res